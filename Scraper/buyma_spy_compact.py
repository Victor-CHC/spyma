import tkinter as tk                
from tkinter import font as tkfont
from tkinter import LEFT, HORIZONTAL, filedialog
import buyma_scraper
import pandas as pd
from openpyxl import Workbook, load_workbook

def dict_list_to_excel(data, filename, extra=False, match=False):

    wb = Workbook()
    
    sheet = wb.active
    sheet.title = 'Items'
    
    # Arrange Column Names

    if len(data) != 0:
        
        if match==False:
            column_names = ['url_ext', 'img', 'item_name', 'item_name_clean', 
                            'sold_amount', 'sold_date']
            column_letters = ['A','B','C','D','E','F']
        else:
            column_names = list(data[0].keys())
            column_letters = []
            for i in range(len(column_names)):
                character = chr(65+i)
                column_letters.append(character)
               
        if extra==True:
            extra_column_names = ['url', 'access_count', 'fav_count', 'syo_id', 'syo_name', 
                                  'syo_img1', 'syo_img_090_1', 'syo_img_210_1', 'syo_url', 
                                  'tanka_format', 'discount_percentage', 'on_timesale', 
                                  'brand_name_eigo', 'brand_url', 'category_id', 'category', 
                                  'buyer_id', 'brand_id', 'model_id', 'season_id', 'thm_id', 
                                  'kokaidate', 'yukodate', 'cate_id1', 'cate_id2', 'cate_id3', 
                                  'tag_ids', 'reference_price_kbn', 'reference_price', 
                                  'timesale_start_date', 'timesale_end_date', 'item_id', 
                                  'price', 'coupon']
            extra_column_letters = ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',
                                    'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X',
                                    'Y', 'Z',
                                    'AA','AB','AC','AD','AE','AF','AG', 'AH', 
                                    'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
                                    'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 
                                    'AW', 'AX', 'AY', 'AZ']
            
            column_names += extra_column_names
            column_letters += extra_column_letters
        
        for i in range(len(column_names)):
            sheet[column_letters[i]+'1'] = column_names[i]
        
        for idx, dat in enumerate(data, start=2):
            for j in range(len(column_letters)):
                
                i = str(idx)
                try:
                    try:
                        sheet[column_letters[j] + i] = dat[column_names[j]]
                    except:
                        sheet[column_letters[j] + i] = str(dat[column_names[j]])
                except:
                    sheet[column_letters[j] + i] = None
    
    wb.save(filename+'.xlsx')



class BuymaSpyCompactApp(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title_font = tkfont.Font(family='Helvetica', size=18, weight="bold", slant="italic")
        self.title('Buyma Spy')

        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        container['bg'] = 'white'

        self.frames = {}
        for F in (SILPage, MISPage, MIFFPage):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("SILPage")

    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()




class SILPage(tk.Frame):

    def sellerListClick():
        input1 = seller_list_url.get()
        input2 = seller_list_prev_days_slider.get()
        
        # If detailed list button is not checked, do a basic scrape
        if details_toggle.get() == 0:
            
            try:
                seller_list_results = buyma_scraper.seller_list(input1, int(input2))
                if len(seller_list_results) == 0:
                    tk.messagebox.showerror("Error","No items found. Try a different URL.")
            except:
                tk.messagebox.showerror("Error", "Unable to get results. Invalid URL or Number.")
                seller_list_results = None
        # If detailed list button is not checked, do a detailed scrape
        else:
    
            try:
                seller_list_results = buyma_scraper.all_listed_items_details(input1, int(input2))
                if len(seller_list_results) == 0:
                    tk.messagebox.showerror("Error","No items found. Try a different URL.")
            except:
                tk.messagebox.showerror("Error", "Unable to get results. Invalid URL or Number.")
                seller_list_results = None
                
        
        #self.config(cursor="")
        
        if seller_list_results:
            directory = filedialog.asksaveasfilename()
            #pd.DataFrame(seller_list_results).to_excel(directory+'.xlsx', index=False)
            if details_toggle.get() == 0:
                dict_list_to_excel(seller_list_results, directory)
            else:
                dict_list_to_excel(seller_list_results, directory, extra=True)
            
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        
        # --- Top Buttons ---
        
        buttonFrame = tk.Frame(self)
        buttonFrame.pack(side="top", fill="x")

        button1 = tk.Button(buttonFrame, text="Sold Items List",
                            command=lambda: controller.show_frame("SILPage"))
        button1['bg'] = '#E72B75'
        button1['fg'] = 'white'
        
        button2 = tk.Button(buttonFrame, text="Matching Items Search",
                            command=lambda: controller.show_frame("MISPage"))
        button2['bg'] = 'white'
        button2['fg'] = '#2AC389'
        
        button3 = tk.Button(buttonFrame, text="Matching Items From File",
                            command=lambda: controller.show_frame("MIFFPage"))    
        button3['bg'] = 'white'
        button3['fg'] = '#4A42B4'
        
        button1.pack(side="left")
        button2.pack(side="left")
        button3.pack(side="left")
        
        
        # --- Seller List ---
        functionFrame = tk.Frame(self)
        functionFrame.pack(side="top", fill="x")        
        
        user_item_frame = tk.LabelFrame(functionFrame, text="Sold Items List")
        user_item_frame['bg'] = 'white'
        user_item_frame['fg'] = '#E72B75'
        user_item_frame.config(font=("Calibri", 18))
        user_item_frame.grid(row=1, column=0, pady=10, padx=20, sticky='n')
        
        seller_list_URL_title = tk.Label(user_item_frame, text="USER'S SALE PAGE URL:")
        seller_list_URL_title['bg'] = 'white'
        seller_list_URL_title['fg'] = '#E72B75'
        seller_list_URL_title.config(font=("Calibri", 12, 'bold'))
        seller_list_URL_title.grid(row=2, column=0, pady=(20,0), padx=20, sticky='w')
        
        global seller_list_url
        seller_list_url = tk.Entry(user_item_frame, width=50)
        seller_list_url.grid(row=3, column=0, pady=(0,20), padx=20, sticky='w')
        seller_list_url['bg'] = 'white'
        seller_list_url.insert(0, 'https://www.buyma.com/buyer/4880785/sales_1.html')
        
        
        seller_list_prev_days_title = tk.Label(user_item_frame, text="PREVIOUS DAYS:")
        seller_list_prev_days_title['bg'] = 'white'
        seller_list_prev_days_title['fg'] = '#E72B75'
        seller_list_prev_days_title.config(font=("Calibri", 12, 'bold'))
        seller_list_prev_days_title.grid(row=4, column=0, pady=0, padx=20, sticky='w')
        
        global seller_list_prev_days_slider
        seller_list_prev_days_slider = tk.Scale(user_item_frame, from_=0, to=100,
                                             tickinterval=20, length=300, orient=HORIZONTAL)
        seller_list_prev_days_slider.set(30)
        seller_list_prev_days_slider['bg'] = 'white'
        seller_list_prev_days_slider['activebackground'] = '#E72B75'
        seller_list_prev_days_slider.grid(row=5, column=0, pady=(0,20), padx=20, sticky='w')
        
        global details_toggle
        details_toggle = tk.IntVar()
        seller_list_details_button = tk.Checkbutton(user_item_frame, text="FULL DETAILS",
                                                 variable=details_toggle, anchor='w')
        seller_list_details_button['bg'] = 'white'
        seller_list_details_button['fg'] = '#E72B75'
        seller_list_details_button.config(font=("Calibri", 12, 'bold'))
        seller_list_details_button.grid(row=6,  pady=5, padx=5, sticky='w')
        
        seller_list_details_caution1 = tk.Label(user_item_frame, 
                                             text="CAUTION: Getting full details may take several minutes.")
        seller_list_details_caution1.config(font=("Calibri", 10, 'italic'))
        
        seller_list_details_caution2 = tk.Label(user_item_frame, 
                                             text="Be patient and wait for the process to finish.")
        seller_list_details_caution2.config(font=("Calibri", 10, 'italic'))
        
        seller_list_details_caution1['bg'] = 'white'
        seller_list_details_caution1['fg'] = 'black'
        seller_list_details_caution2['bg'] = 'white'
        seller_list_details_caution2['fg'] = 'black'
        seller_list_details_caution1.grid(row=7, column=0, pady=0, padx=20, sticky='w')
        seller_list_details_caution2.grid(row=8, column=0, pady=0, padx=20, sticky='w')
        
        
        sellerListButton = tk.Button(user_item_frame, text="Download Item List", 
                                  command=SILPage.sellerListClick)
        sellerListButton['bg'] = 'white'
        sellerListButton['activebackground'] = '#E72B75'
        sellerListButton['activeforeground'] = 'white'
        sellerListButton.grid(row=10, column=0, pady=30, padx=20, sticky='w')
        
        
        
        


class MISPage(tk.Frame):

    def itemMatchListClick():
        input1 = fuzzy_item_query.get()
        input2 = fuzzy_threshold_slider.get()
      
        try:
            item_list_results = buyma_scraper.similar_extra_search(input1, int(input2))
            item_count = len(item_list_results)
            if item_count==0:
                tk.messagebox.showerror("Error","No similar items found. Try a different search query or lower the threshold.")
        except:
            tk.messagebox.showerror("Error", "Invalid Search Word")
            item_list_results = None
    
        if item_list_results:
            directory = filedialog.asksaveasfilename()
            #pd.DataFrame(item_list_results).to_excel(directory+'.xlsx', index=False)
            dict_list_to_excel(item_list_results, directory, match=True)    

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        
        # --- Top Buttons ---
        
        buttonFrame = tk.Frame(self)
        buttonFrame.pack(side="top", fill="x")

        button1 = tk.Button(buttonFrame, text="Sold Items List",
                            command=lambda: controller.show_frame("SILPage"))
        button1['bg'] = 'white'
        button1['fg'] = '#E72B75'
        
        button2 = tk.Button(buttonFrame, text="Matching Items Search",
                            command=lambda: controller.show_frame("MISPage"))
        button2['bg'] = '#2AC389'
        button2['fg'] = 'white'
        
        button3 = tk.Button(buttonFrame, text="Matching Items From File",
                            command=lambda: controller.show_frame("MIFFPage"))    
        button3['bg'] = 'white'
        button3['fg'] = '#4A42B4'  
        
        button1.pack(side="left")
        button2.pack(side="left")
        button3.pack(side="left")
        
        # --- Item Matching ---
        functionFrame = tk.Frame(self)
        functionFrame.pack(side="top", fill="x") 
        
        item_matching_frame = tk.LabelFrame(functionFrame, text="Matching Items Search")
        item_matching_frame['bg'] = 'white'
        item_matching_frame['fg'] = '#2AC389'
        item_matching_frame.config(font=("Calibri", 18))
        item_matching_frame.grid(row=1, column=1, pady=10, padx=20, sticky='n')
        
        fuzzy_item_queryL_title = tk.Label(item_matching_frame, text="SEARCH WORDS:")
        fuzzy_item_queryL_title['bg'] = 'white'
        fuzzy_item_queryL_title['fg'] = '#2AC389'
        fuzzy_item_queryL_title.config(font=("Calibri", 12, 'bold'))
        fuzzy_item_queryL_title.grid(row=2, column=0, pady=(20,0), padx=20, sticky='w')
        
        global fuzzy_item_query
        fuzzy_item_query = tk.Entry(item_matching_frame, width=50)
        fuzzy_item_query.grid(row=3, column=0, pady=(0,20), padx=20, sticky='w')
        fuzzy_item_query['bg'] = 'white'
        fuzzy_item_query.insert(0, 'Air Jordan')
        
        
        fuzzy_threshold_slider_title = tk.Label(item_matching_frame, text="WORD MATCH THRESHOLD")
        fuzzy_threshold_slider_title['bg'] = 'white'
        fuzzy_threshold_slider_title['fg'] = '#2AC389'
        fuzzy_threshold_slider_title.config(font=("Calibri", 12, 'bold'))
        fuzzy_threshold_slider_title.grid(row=4, column=0, pady=0, padx=20, sticky='w')
        
        global fuzzy_threshold_slider
        fuzzy_threshold_slider = tk.Scale(item_matching_frame, from_=0, to=100,
                                             tickinterval=20, length=300, orient=HORIZONTAL)
        fuzzy_threshold_slider.set(90)
        fuzzy_threshold_slider['bg'] = 'white'
        fuzzy_threshold_slider['activebackground'] = '#2AC389'
        fuzzy_threshold_slider.grid(row=5, column=0, pady=(0,20), padx=20, sticky='w')
        
        
        itemMatchListButton = tk.Button(item_matching_frame, text="Download Item List", 
                                  command=MISPage.itemMatchListClick)
        itemMatchListButton['bg'] = 'white'
        itemMatchListButton['activebackground'] = '#2AC389'
        itemMatchListButton['activeforeground'] = 'white'
        itemMatchListButton.grid(row=10, column=0, pady=30, padx=20, sticky='w')

class MIFFPage(tk.Frame):

    def importFileClick():
        global import_list
        
        filename = filedialog.askopenfilename(filetypes=[('Excel', ('*.xls', '*.xlsx'))])
        wb = load_workbook(filename)
        
        sheetname = wb.sheetnames[0]
        sheetdata = wb[sheetname].values
        columns = next(sheetdata)[0:]
        
        df = pd.DataFrame(sheetdata, columns=columns)
        
        
        #df = pd.read_excel(filename)
        df_dict = df.T.to_dict()
        import_list = list(df_dict.values())
        
        # Print the file name
        filename_label = tk.Label(matching_user_item_frame, text=filename)    
        filename_label.grid(row=4, column=0, padx=20, sticky='w')
        filename_label['bg'] = 'white'
        filename_label['fg'] = '#4A42B4'
    
    
    
    def fileMatchListClick():
        
        threshold = excel_threshold_slider.get()
        
        try:
            results = buyma_scraper.similar_items(import_list, threshold)
        except:
            results = None
            tk.messagebox.showerror("Error", "Invalid File")
        
        if results:
            directory = filedialog.asksaveasfilename()
            pd.DataFrame(results).to_excel(directory+'.xlsx', index=False)    

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        
        # --- Top Buttons ---
        
        buttonFrame = tk.Frame(self)
        buttonFrame.pack(side="top", fill="x")

        button1 = tk.Button(buttonFrame, text="Sold Items List",
                            command=lambda: controller.show_frame("SILPage"))
        button1['bg'] = 'white'
        button1['fg'] = '#E72B75'
        
        button2 = tk.Button(buttonFrame, text="Matching Items Search",
                            command=lambda: controller.show_frame("MISPage"))
        button2['bg'] = 'white'
        button2['fg'] = '#2AC389'
        
        button3 = tk.Button(buttonFrame, text="Matching Items From File",
                            command=lambda: controller.show_frame("MIFFPage"))    
        button3['bg'] = '#4A42B4'
        button3['fg'] = 'white'
        
        button1.pack(side="left")
        button2.pack(side="left")
        button3.pack(side="left")
        
        
        # --- Item Matching From List---
        functionFrame = tk.Frame(self)
        functionFrame.pack(side="top", fill="x") 
        
        global matching_user_item_frame
        matching_user_item_frame = tk.LabelFrame(functionFrame, text="Matching Items From File")
        matching_user_item_frame['bg'] = 'white'
        matching_user_item_frame['fg'] = '#4A42B4'
        matching_user_item_frame.config(font=("Calibri", 18))
        matching_user_item_frame.grid(row=1, column=2, pady=10, padx=20, sticky='n')
        
        
        import_excel_title = tk.Label(matching_user_item_frame, text="IMPORT EXCEL FILE:")
        import_excel_title['bg'] = 'white'
        import_excel_title['fg'] = '#4A42B4'
        import_excel_title.config(font=("Calibri", 12, 'bold'))
        import_excel_title.grid(row=2, column=0, pady=(20,0), padx=20, sticky='w')
        
        open_file_btn = tk.Button(matching_user_item_frame, text="Open File", 
                               command=MIFFPage.importFileClick)
        open_file_btn['bg'] = 'white'
        open_file_btn['activebackground'] = '#4A42B4'
        open_file_btn['activeforeground'] = 'white'
        open_file_btn.grid(row=3, column=0, pady=(0,0), padx=20, sticky='w')
        
        excel_threshold_slider_title = tk.Label(matching_user_item_frame, text="WORD MATCH THRESHOLD")
        excel_threshold_slider_title['bg'] = 'white'
        excel_threshold_slider_title['fg'] = '#4A42B4'
        excel_threshold_slider_title.config(font=("Calibri", 12, 'bold'))
        excel_threshold_slider_title.grid(row=5, column=0, pady=(20,0), padx=20, sticky='w')
        
        global excel_threshold_slider
        excel_threshold_slider = tk.Scale(matching_user_item_frame, from_=0, to=100,
                                             tickinterval=20, length=300, orient=HORIZONTAL)
        excel_threshold_slider.set(90)
        excel_threshold_slider['bg'] = 'white'
        excel_threshold_slider['activebackground'] = '#4A42B4'
        excel_threshold_slider.grid(row=6, column=0, pady=(0,20), padx=20, sticky='w')
        
        
        fileMatchListButton = tk.Button(matching_user_item_frame, text="Download Item List", 
                                  command=MIFFPage.fileMatchListClick)
        fileMatchListButton['bg'] = 'white'
        fileMatchListButton['activebackground'] = '#4A42B4'
        fileMatchListButton['activeforeground'] = 'white'
        
        
        fileMatchListButton.grid(row=7, column=0, pady=30, padx=20, sticky='w')
        


if __name__ == "__main__":
    app = BuymaSpyCompactApp()
    app.mainloop()