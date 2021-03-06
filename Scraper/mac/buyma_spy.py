from tkinter import *
from tkinter import filedialog
import buyma_scraper
import pandas as pd
from openpyxl import Workbook, load_workbook

def dict_list_to_excel(data, filename, extra=False, match=False):

    wb = Workbook()
    
    sheet = wb.active
    sheet.title = 'Items'
    
    # Arrange Column Names

    if len(data) != 0:
        
        if match==False:
            column_names = ['url_ext', 'img', 'item_name', 'item_name_clean', 
                            'sold_amount', 'sold_date']
            column_letters = ['A','B','C','D','E','F']
        else:
            column_names = list(data[0].keys())
            column_letters = []
            for i in range(len(column_names)):
                character = chr(65+i)
                column_letters.append(character)
               
        if extra==True:
            extra_column_names = ['url', 'access_count', 'fav_count', 'syo_id', 'syo_name', 
                                  'syo_img1', 'syo_img_090_1', 'syo_img_210_1', 'syo_url', 
                                  'tanka_format', 'discount_percentage', 'on_timesale', 
                                  'brand_name_eigo', 'brand_url', 'category_id', 'category', 
                                  'buyer_id', 'brand_id', 'model_id', 'season_id', 'thm_id', 
                                  'kokaidate', 'yukodate', 'cate_id1', 'cate_id2', 'cate_id3', 
                                  'tag_ids', 'reference_price_kbn', 'reference_price', 
                                  'timesale_start_date', 'timesale_end_date', 'item_id', 
                                  'price', 'coupon']
            extra_column_letters = ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',
                                    'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X',
                                    'Y', 'Z',
                                    'AA','AB','AC','AD','AE','AF','AG', 'AH', 
                                    'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
                                    'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 
                                    'AW', 'AX', 'AY', 'AZ']
            
            column_names += extra_column_names
            column_letters += extra_column_letters
        
        for i in range(len(column_names)):
            sheet[column_letters[i]+'1'] = column_names[i]
        
        for idx, dat in enumerate(data, start=2):
            for j in range(len(column_letters)):
                
                i = str(idx)
                try:
                    try:
                        sheet[column_letters[j] + i] = dat[column_names[j]]
                    except:
                        sheet[column_letters[j] + i] = str(dat[column_names[j]])
                except:
                    sheet[column_letters[j] + i] = None
    
    wb.save(filename+'.xlsx')
    

def sellerListClick():
    input1 = seller_list_url.get()
    input2 = seller_list_prev_days_slider.get()
    
    # If detailed list button is not checked, do a basic scrape
    if details_toggle.get() == 0:
        
        try:
            seller_list_results = buyma_scraper.seller_list(input1, int(input2))
            if len(seller_list_results) == 0:
                messagebox.showerror("Error","No items found. Try a different URL.")
        except:
            messagebox.showerror("Error", "Unable to get results. Invalid URL or Number.")
            seller_list_results = None
    # If detailed list button is not checked, do a detailed scrape
    else:

        try:
            seller_list_results = buyma_scraper.all_listed_items_details(input1, int(input2))
            if len(seller_list_results) == 0:
                messagebox.showerror("Error","No items found. Try a different URL.")
        except:
            messagebox.showerror("Error", "Unable to get results. Invalid URL or Number.")
            seller_list_results = None
            
    
    root.config(cursor="")
    
    if seller_list_results:
        directory = filedialog.asksaveasfilename()
        #pd.DataFrame(seller_list_results).to_excel(directory+'.xlsx', index=False)
        if details_toggle.get() == 0:
            dict_list_to_excel(seller_list_results, directory)
        else:
            dict_list_to_excel(seller_list_results, directory, extra=True)

        
def itemMatchListClick():
    input1 = fuzzy_item_query.get()
    input2 = fuzzy_threshold_slider.get()
  
    try:
        item_list_results = buyma_scraper.similar_extra_search(input1, int(input2))
        item_count = len(item_list_results)
        if item_count==0:
            messagebox.showerror("Error","No similar items found. Try a different search query or lower the threshold.")
    except:
        messagebox.showerror("Error", "Invalid URL")
        item_list_results = None

    if item_list_results:
        directory = filedialog.asksaveasfilename()
        #pd.DataFrame(item_list_results).to_excel(directory+'.xlsx', index=False)
        dict_list_to_excel(item_list_results, directory, match=True)


def importFileClick():
    global import_list
    
    filename = filedialog.askopenfilename(filetypes=[('Excel', ('*.xls', '*.xlsx'))])
    wb = load_workbook(filename)
    
    sheetname = wb.sheetnames[0]
    sheetdata = wb[sheetname].values
    columns = next(sheetdata)[0:]
    
    df = pd.DataFrame(sheetdata, columns=columns)
    
    
    #df = pd.read_excel(filename)
    df_dict = df.T.to_dict()
    import_list = list(df_dict.values())
    
    # Print the file name
    filename_label = Label(matching_user_item_frame, text=filename)    
    filename_label.grid(row=4, column=0, padx=20, sticky='w')
    filename_label['bg'] = 'white'
    filename_label['fg'] = '#4A42B4'



def fileMatchListClick():
    
    threshold = excel_threshold_slider.get()
    
    try:
        results = buyma_scraper.similar_items(import_list, threshold)
    except:
        results = None
        messagebox.showerror("Error", "Invalid File")
    
    if results:
        directory = filedialog.asksaveasfilename()
        pd.DataFrame(results).to_excel(directory+'.xlsx', index=False)
        


root = Tk()

# --- App Base Details ---
root['bg'] = 'white'
root.title('Buyma Spy')
root.iconbitmap("icon_purple.icns")


# buyma_title_img = PhotoImage(file='logo_purple.png')
# buyma_title_label = Label(image=buyma_title_img,
#                        borderwidth=0)
# buyma_title_label.grid(row=0, column=0, pady=(10,20), padx=20, sticky='w')



# --- Seller List ---
user_item_frame = LabelFrame(root, text="Sold Items List")
user_item_frame['bg'] = 'white'
user_item_frame['fg'] = '#E72B75'
user_item_frame.config(font=("Calibri", 18))
user_item_frame.grid(row=1, column=0, pady=10, padx=20, sticky='n')

seller_list_URL_title = Label(user_item_frame, text="USER'S SALE PAGE URL:")
seller_list_URL_title['bg'] = 'white'
seller_list_URL_title['fg'] = '#E72B75'
seller_list_URL_title.config(font=("Calibri", 12, 'bold'))
seller_list_URL_title.grid(row=2, column=0, pady=(20,0), padx=20, sticky='w')


seller_list_url = Entry(user_item_frame, width=38)
seller_list_url.grid(row=3, column=0, pady=(0,20), padx=20, sticky='w')
seller_list_url['bg'] = 'white'
seller_list_url.insert(0, 'https://www.buyma.com/buyer/4880785/sales_1.html')


seller_list_prev_days_title = Label(user_item_frame, text="PREVIOUS DAYS:")
seller_list_prev_days_title['bg'] = 'white'
seller_list_prev_days_title['fg'] = '#E72B75'
seller_list_prev_days_title.config(font=("Calibri", 12, 'bold'))
seller_list_prev_days_title.grid(row=4, column=0, pady=0, padx=20, sticky='w')

seller_list_prev_days_slider = Scale(user_item_frame, from_=0, to=100,
                                     tickinterval=20, length=280, orient=HORIZONTAL)
seller_list_prev_days_slider.set(30)
seller_list_prev_days_slider['bg'] = 'white'
seller_list_prev_days_slider['activebackground'] = '#E72B75'
seller_list_prev_days_slider.grid(row=5, column=0, pady=(0,20), padx=20, sticky='w')

details_toggle = IntVar()
seller_list_details_button = Checkbutton(user_item_frame, text="FULL DETAILS",
                                         variable=details_toggle, anchor='w')
seller_list_details_button['bg'] = 'white'
seller_list_details_button['fg'] = '#E72B75'
seller_list_details_button.config(font=("Calibri", 12, 'bold'))
seller_list_details_button.grid(row=6,  pady=5, padx=5, sticky='w')

seller_list_details_caution1 = Label(user_item_frame, 
                                     text="CAUTION: Getting full details may take several minutes.")
seller_list_details_caution1.config(font=("Calibri", 11, 'italic'))

seller_list_details_caution2 = Label(user_item_frame, 
                                     text="Be patient and wait for the process to finish.")
seller_list_details_caution2.config(font=("Calibri", 11, 'italic'))

seller_list_details_caution1['bg'] = 'white'
seller_list_details_caution1['fg'] = 'black'
seller_list_details_caution2['bg'] = 'white'
seller_list_details_caution2['fg'] = 'black'
seller_list_details_caution1.grid(row=7, column=0, pady=0, padx=20, sticky='w')
seller_list_details_caution2.grid(row=8, column=0, pady=0, padx=20, sticky='w')


sellerListButton = Button(user_item_frame, text="Download Item List", 
                          command=sellerListClick)
sellerListButton['bg'] = 'white'
sellerListButton['activebackground'] = '#E72B75'
sellerListButton['activeforeground'] = 'white'
sellerListButton.grid(row=10, column=0, pady=30, padx=20, sticky='w')




# --- Item Matching ---

item_matching_frame = LabelFrame(root, text="Matching Items Search")
item_matching_frame['bg'] = 'white'
item_matching_frame['fg'] = '#2AC389'
item_matching_frame.config(font=("Calibri", 18))
item_matching_frame.grid(row=1, column=1, pady=10, padx=20, sticky='n')

fuzzy_item_queryL_title = Label(item_matching_frame, text="SEARCH WORDS:")
fuzzy_item_queryL_title['bg'] = 'white'
fuzzy_item_queryL_title['fg'] = '#2AC389'
fuzzy_item_queryL_title.config(font=("Calibri", 12, 'bold'))
fuzzy_item_queryL_title.grid(row=2, column=0, pady=(20,0), padx=20, sticky='w')


fuzzy_item_query = Entry(item_matching_frame, width=30)
fuzzy_item_query.grid(row=3, column=0, pady=(0,20), padx=20, sticky='w')
fuzzy_item_query['bg'] = 'white'
fuzzy_item_query.insert(0, 'Air Jordan')


fuzzy_threshold_slider_title = Label(item_matching_frame, text="WORD MATCH THRESHOLD")
fuzzy_threshold_slider_title['bg'] = 'white'
fuzzy_threshold_slider_title['fg'] = '#2AC389'
fuzzy_threshold_slider_title.config(font=("Calibri", 12, 'bold'))
fuzzy_threshold_slider_title.grid(row=4, column=0, pady=0, padx=20, sticky='w')

fuzzy_threshold_slider = Scale(item_matching_frame, from_=0, to=100,
                                     tickinterval=20, length=280, orient=HORIZONTAL)
fuzzy_threshold_slider.set(90)
fuzzy_threshold_slider['bg'] = 'white'
fuzzy_threshold_slider['activebackground'] = '#2AC389'
fuzzy_threshold_slider.grid(row=5, column=0, pady=(0,20), padx=20, sticky='w')


itemMatchListButton = Button(item_matching_frame, text="Download Item List", 
                          command=itemMatchListClick)
itemMatchListButton['bg'] = 'white'
itemMatchListButton['activebackground'] = '#2AC389'
itemMatchListButton['activeforeground'] = 'white'
itemMatchListButton.grid(row=10, column=0, pady=30, padx=20, sticky='w')


# --- Item Matching From List---
matching_user_item_frame = LabelFrame(root, text="Matching Items From File")
matching_user_item_frame['bg'] = 'white'
matching_user_item_frame['fg'] = '#4A42B4'
matching_user_item_frame.config(font=("Calibri", 18))
matching_user_item_frame.grid(row=1, column=2, pady=10, padx=20, sticky='n')


import_excel_title = Label(matching_user_item_frame, text="IMPORT EXCEL FILE:")
import_excel_title['bg'] = 'white'
import_excel_title['fg'] = '#4A42B4'
import_excel_title.config(font=("Calibri", 12, 'bold'))
import_excel_title.grid(row=2, column=0, pady=(20,0), padx=20, sticky='w')

open_file_btn = Button(matching_user_item_frame, text="Open File", 
                       command=importFileClick)
open_file_btn['bg'] = 'white'
open_file_btn['activebackground'] = '#4A42B4'
open_file_btn['activeforeground'] = 'white'
open_file_btn.grid(row=3, column=0, pady=(0,0), padx=20, sticky='w')

excel_threshold_slider_title = Label(matching_user_item_frame, text="WORD MATCH THRESHOLD")
excel_threshold_slider_title['bg'] = 'white'
excel_threshold_slider_title['fg'] = '#4A42B4'
excel_threshold_slider_title.config(font=("Calibri", 12, 'bold'))
excel_threshold_slider_title.grid(row=5, column=0, pady=(20,0), padx=20, sticky='w')

excel_threshold_slider = Scale(matching_user_item_frame, from_=0, to=100,
                                     tickinterval=20, length=300, orient=HORIZONTAL)
excel_threshold_slider.set(90)
excel_threshold_slider['bg'] = 'white'
excel_threshold_slider['activebackground'] = '#4A42B4'
excel_threshold_slider.grid(row=6, column=0, pady=(0,20), padx=20, sticky='w')


fileMatchListButton = Button(matching_user_item_frame, text="Download Item List", 
                          command=fileMatchListClick)
fileMatchListButton['bg'] = 'white'
fileMatchListButton['activebackground'] = '#4A42B4'
fileMatchListButton['activeforeground'] = 'white'


fileMatchListButton.grid(row=7, column=0, pady=30, padx=20, sticky='w')


root.mainloop()